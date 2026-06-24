"""
Generate 14 department Excel sheets for offline KPI data collection.

Columns match the portal entry form:
  KPI name · Field 1 label · Field 2 label · blank cells for values

Labels from backend/constants/kpiFieldLabels.json (same as the app).

Run from repo root:
  python docs/generate_dept_excel_sheets.py

Output:
  docs/excel/D01_Health_NHM.xlsx … D14_DC_Office.xlsx
  docs/excel/Tirap_All_Depts_KPI_Collection.xlsx
  docs/Tirap_KPI_Data_Collection_Template.csv
"""
import csv
import json
import re
from copy import copy
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "backend" / "constants" / "kpiCatalog.json"
FIELD_LABELS = ROOT / "backend" / "constants" / "kpiFieldLabels.json"
OUT_DIR = Path(__file__).resolve().parent / "excel"
CSV_OUT = Path(__file__).resolve().parent / "Tirap_KPI_Data_Collection_Template.csv"
OUT_DIR.mkdir(parents=True, exist_ok=True)

DEPARTMENTS = [
    ("D01", "Health & NHM", "D01_Health_NHM"),
    ("D02", "School Education", "D02_School_Education"),
    ("D03", "ICDS / Women & Child Development", "D03_ICDS_WCD"),
    ("D04", "Social Justice / De-addiction / Youth", "D04_Social_Justice"),
    ("D05", "Rural Sanitation / Urban Local Bodies", "D05_Sanitation_ULB"),
    ("D06", "PWD / PMGSY / Water-Power Infrastructure", "D06_PWD_Infra"),
    ("D07", "Power Department / DISCOM", "D07_Power_DISCOM"),
    ("D08", "Police / Law & Order", "D08_Police"),
    ("D09", "Agriculture / Allied Livelihoods", "D09_Agriculture"),
    ("D10", "Food & Civil Supplies / Welfare", "D10_Food_Welfare"),
    ("D11", "Revenue / District Administration", "D11_Revenue"),
    ("D12", "DLR / Transport / Citizen Services", "D12_Transport_DLR"),
    ("D13", "Disaster Management", "D13_Disaster"),
    ("D14", "District Governance / DC Office", "D14_DC_Office"),
]

HEADERS = [
    "S.No",
    "KPI Indicator",
    "Reporting",
    "Field 1",
    "Field 2",
    "Reporting Month",
    "Value for Field 1",
    "Value for Field 2",
    "Yes/No (1 or 0)",
    "Officer Name",
    "Designation",
    "Phone",
    "Remarks",
]

COL_MONTH = 6
COL_VAL1 = 7
COL_VAL2 = 8
COL_YESNO = 9
INPUT_COLS = {COL_MONTH, COL_VAL1, COL_VAL2, COL_YESNO, 10, 11, 12, 13}

FREQ = {"M": "Monthly", "Q": "Quarterly"}
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
NOTE_FONT = Font(size=10, color="5C6B7A")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ROW_ALT = PatternFill("solid", fgColor="F8FAFC")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E6")


def make_kpi_id(dept_id, name):
    """Match backend/constants/kpiCatalog.js makeKpiId()."""
    s = str(name).upper()
    for old, new in [
        (" ", "_"), ("/", "_"), ("-", "_"), ("(", ""), (")", ""),
        ("%", "PCT"), ("&", "AND"), (".", ""), (",", ""),
        ("+", "PLUS"), (">", ""), ("<", ""), ("=", ""),
        ("≥", ""), ("'", ""), ("×", "X"),
    ]:
        s = s.replace(old, new)
    return f"{dept_id}_{s}"


def is_yes_no(kpi):
    return (kpi.get("unitLabel") or "").lower() == "yes/no"


def period_label(kpi):
    freq = FREQ.get(kpi.get("freq", "M"), "Monthly")
    if is_yes_no(kpi):
        return f"{freq} (Yes/No)"
    unit = kpi.get("unitLabel") or kpi.get("unit") or ""
    if unit in ("%", "pct"):
        return f"{freq} (%)"
    if kpi.get("unit") == "rate" or "per lakh" in kpi.get("name", "").lower():
        return f"{freq} (per lakh)"
    return freq


def resolve_field_labels(kpi, labels_map):
    if is_yes_no(kpi):
        return "—", "—"
    kpi_id = make_kpi_id(kpi["deptId"], kpi["name"])
    pair = labels_map.get(kpi_id)
    if pair and len(pair) == 2:
        return pair[0], pair[1]
    if kpi.get("polarity") == "LOWER":
        return "Total base", "Count"
    return "Field 1", "Field 2"


def style_data_sheet(ws, dept_id, dept_name, kpis, labels_map):
    ws.merge_cells("A1:M1")
    ws["A1"] = f"Tirap District — {dept_id}: {dept_name}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        "Fill yellow cells only. Enter the reporting month (e.g. 2026-05) and the two counts "
        "shown in Field 1 / Field 2 columns. For Yes/No indicators use column I (1 = Yes, 0 = No)."
    )
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)

    header_row = 4
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, kpi in enumerate(kpis):
        row = header_row + 1 + i
        yes_no = is_yes_no(kpi)
        f1, f2 = resolve_field_labels(kpi, labels_map)
        values = [
            kpi["num"],
            kpi["name"],
            period_label(kpi),
            f1,
            f2,
            "",
            "" if yes_no else "",
            "" if yes_no else "",
            "" if yes_no else "—",
            "",
            "",
            "",
            "",
        ]
        if yes_no:
            values[6] = "—"
            values[7] = "—"

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 1:
                cell.fill = ROW_ALT
            if col in INPUT_COLS:
                if yes_no and col in (COL_VAL1, COL_VAL2):
                    pass
                elif yes_no and col == COL_YESNO:
                    cell.fill = INPUT_FILL
                elif not yes_no and col in (COL_MONTH, COL_VAL1, COL_VAL2, 10, 11, 12, 13):
                    cell.fill = INPUT_FILL
                elif not yes_no and col == COL_YESNO:
                    pass

    widths = [6, 44, 14, 28, 28, 14, 14, 14, 12, 18, 16, 14, 22]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[header_row].height = 28


def build_workbook(dept_id, dept_name, kpis, labels_map):
    wb = Workbook()
    ws = wb.active
    safe = re.sub(r"[^\w\-]", "_", dept_name)[:24]
    ws.title = f"{dept_id}_{safe}"[:31]
    style_data_sheet(ws, dept_id, dept_name, kpis, labels_map)
    return wb


def copy_sheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font) if cell.font else Font()
                new_cell.fill = copy(cell.fill) if cell.fill else PatternFill()
                new_cell.border = copy(cell.border) if cell.border else Border()
                new_cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
    for col, dim in source_ws.column_dimensions.items():
        if dim.width:
            target_ws.column_dimensions[col].width = dim.width
    target_ws.freeze_panes = source_ws.freeze_panes
    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))


def write_csv(rows):
    fieldnames = [
        "Dept ID",
        "Department",
        "S.No",
        "KPI Indicator",
        "Reporting",
        "Field 1",
        "Field 2",
        "Reporting Month",
        "Value for Field 1",
        "Value for Field 2",
        "Yes/No (1 or 0)",
        "Officer Name",
        "Designation",
        "Phone",
        "Remarks",
    ]
    with CSV_OUT.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    labels_map = json.loads(FIELD_LABELS.read_text(encoding="utf-8"))
    all_kpis = catalog["kpis"]

    combined = Workbook()
    combined.remove(combined.active)
    csv_rows = []

    for dept_id, dept_name, file_stub in DEPARTMENTS:
        kpis = sorted(
            [k for k in all_kpis if k["deptId"] == dept_id],
            key=lambda k: k["num"],
        )
        for k in kpis:
            k["deptId"] = dept_id

        wb = build_workbook(dept_id, dept_name, kpis, labels_map)
        out_path = OUT_DIR / f"{file_stub}.xlsx"
        wb.save(out_path)
        print(f"  {out_path.name} ({len(kpis)} indicators)")

        ws = wb.active
        safe = re.sub(r"[^\w\-]", "_", dept_name)[:18]
        new_ws = combined.create_sheet(title=f"{dept_id}_{safe}"[:31])
        copy_sheet(ws, new_ws)

        for kpi in kpis:
            yes_no = is_yes_no(kpi)
            f1, f2 = resolve_field_labels(kpi, labels_map)
            csv_rows.append({
                "Dept ID": dept_id,
                "Department": dept_name,
                "S.No": kpi["num"],
                "KPI Indicator": kpi["name"],
                "Reporting": period_label(kpi),
                "Field 1": "—" if yes_no else f1,
                "Field 2": "—" if yes_no else f2,
                "Reporting Month": "",
                "Value for Field 1": "",
                "Value for Field 2": "",
                "Yes/No (1 or 0)": "" if not yes_no else "",
                "Officer Name": "",
                "Designation": "",
                "Phone": "",
                "Remarks": "",
            })

    combined_path = OUT_DIR / "Tirap_All_Depts_KPI_Collection.xlsx"
    combined.save(combined_path)
    print(f"  {combined_path.name} (14 sheets)")

    write_csv(csv_rows)
    print(f"  {CSV_OUT.name} ({len(csv_rows)} rows)")
    print(f"\nDone — files in {OUT_DIR}")


if __name__ == "__main__":
    main()
